from models import *
import falcon
from jinjatemp import *
import uuid
import hashlib
import openpyxl
import os
UPLOAD_FOLDER= os.getcwd()


Welcome_Template = templateEnv.get_template('welcome.html')
reg_template=templateEnv.get_template('reg.html')
login_template=templateEnv.get_template('login.html')
bookmark1_template=templateEnv.get_template('bookmark1.html')
bookmark_form=templateEnv.get_template('bookmark_form.html')
bookmark_form2=templateEnv.get_template('bookmark_form2.html')
bookmark_view=templateEnv.get_template('bookmark_view.html')
bookmark_view2=templateEnv.get_template('bookmark_view2.html')
upload_file='upload.html'
upload_file=templateEnv.get_template(upload_file)

def _hash_password(password):
    print "password", password
    pass_id= uuid.uuid4().hex
    password= hashlib.sha256(pass_id.encode() + password.encode()).hexdigest() + ":" + pass_id
    return password

def _check_password(hashed_password, user_password):
    print "hash",hashed_password
    password, pass_id = hashed_password.split(':')
    return password == hashlib.sha256(pass_id.encode()+ user_password.encode()).hexdigest()


class firstPage(object):
    def on_get(self, req, resp):
        resp.content_type = "text/html"
        print "I am here"
        resp.body = Welcome_Template.render()
        resp.status = falcon.HTTP_301
        print resp.status

class registration(object):
    def on_get(self, req, resp):
        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200 #200 is for ok ,if we dont give status it will by default took it """
#as 200, but it is necessary to give status everytime"""
        resp.body = reg_template.render()

    def on_post(self, req, resp):
        print "Reached Here"
        details= req.params
        print details

        templateVars = ""
        try:
            print "reached in try"
            #print details.get('Email')
            #print registration.objects.get(email_db = details.get('Email'))
            check =  registration_db.objects.get(email_db = details.get('Email'))
            templateVars = "UserName already exists"
        except DoesNotExist:
            # save new user here
            uid=str(uuid.uuid4())
            new_user = registration_db(u_id=uid,name_db=details.get('Name') , email_db=details.get('Email'))
            print new_user.email_db
            print details.get('Password')
            new_user.create_password = _hash_password(details.get('Password'))
            if (details.get('Password')==details.get('Confirm_Password')):
                print '1'
                new_user.save()
                templateVars="Registered successfully"
                resp.content_type= "text/html"

                resp.set_header('Location','/bookmark1/'+uid)
                resp.status=falcon.HTTP_301
                return

            else:
                print '2'
                templateVars = "password mismatch "
        except Exception, e:
            print '3'
            templateVars= "Database error, Redirect to LOG IN Page"
            print str(e)

        resp.content_type= "text/html"
        resp.body = reg_template.render(jsob = templateVars)
        resp.status=falcon.HTTP_200




class login(object):
    def on_get(self, req, resp):

        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        resp.body = login_template.render()

    def on_post(self,req,resp):
        details= req.params
        print "login wale",details
        resp.content_type = "text/html"
        try:
            if (registration_db.objects.get(email_db=details.get('EMAILID')) ):
                print '1'
                try:
                    print '2'
                    d=registration_db.objects.get(email_db=details.get('EMAILID'))
                    hashed_password=d.create_password
                    user_password=details.get('PASSWORD')
                    if(_check_password(hashed_password, user_password)):

                        print 'successfully logged in'
                        #x_auth = uuid.uuid4().get_hex()
                        #print type(x_auth)
                        #print x_auth
                        #d.x_auth=x_auth
                        d.save()
                        try:
                            #resp.set_cookie("mycookie",x_auth,expires=None, max_age=None, domain=None, path=None, secure=False, http_only=False)
                            #val = req.cookies
                            #print val
                            #print val["mycookie"]
                            #print type(val)
                            uid=d.u_id
                            resp.set_header('Location','/bookmark1/'+uid)
                            resp.status=falcon.HTTP_301
                            return
                            
                        except Exception ,e:
                            print str(e)

                    else:
                        temp= 'passwrd mismatch'
                except DoesNotExist:
                    print 'heloo'
                    return
        except DoesNotExist:
            temp="email id not Registered"
        resp.body= login_template.render(jsob=temp)
        resp.status=falcon.HTTP_200

class bookmark1(object):
    def on_get(self, req, resp,uid):

        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        
        resp.body = bookmark1_template.render(jsob=uid)
class add_bookmark(object):
    def on_get(self, req, resp,uid):
        print"rdgr", uid
        resp.content_type = "text/html"
        
        resp.body = bookmark_form.render()
    def on_post(self,req,resp,uid):
        

        details= req.params
        print details
        #print details.URL
        
        try:
            if Bookmark_db.objects.get(location=details.get('URL'),u_id2=uid):
                #resp.set_header('Location','/add_bookmark/uid')
                mssg="url already present"
                resp.content_type = "text/html"
        
                resp.body = bookmark_form.render(jsob=uid,mssg=mssg)
                resp.status=falcon.HTTP_301
        except DoesNotExist:
            bookmark_db=Bookmark_db(name=details.get('Name'),location=details.get('URL'),labels=details.get('Label'),notes=details.get('Notes'),u_id2=uid).save()
            print "successfully added"
            resp.set_header('Location','/views/'+uid)
            resp.status=falcon.HTTP_301
            return

class views(object):
    def on_get(self, req, resp,uid):

        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        list1=[]
       
        if not Bookmark_db.objects(u_id2=uid):
            print "1"
            temp="No Bookmark available"
            resp.body = bookmark_view.render(jsob=uid,temp=temp)
            resp.status=falcon.HTTP_301
            return
        for i in Bookmark_db.objects(u_id2=uid):
            list1.append(i.location)
        print "views wale ", list1
        

        
        resp.body = bookmark_view.render(url_list=list1,jsob=uid)
        resp.status=falcon.HTTP_301
class remove(object):
    def on_get(self, req, resp,uid,aa):
        print "remove wala" ,aa
        c=Bookmark_db.objects.get(u_id2=uid,location="http://"+aa)
        c.delete()
        temp="file deleted successfully"
        resp.content_type = "text/html"
        resp.set_header('Location','/views/'+uid)
        resp.status=falcon.HTTP_301
        return
class edit(object):
    def on_get(self, req, resp,uid,aa):
        print "url",aa
        print "http://"+aa
        print uid
        c=Bookmark_db.objects.get(u_id2=uid,location="http://"+ aa)
        d= c.name
        #e=c.location
        f=c.labels

        g=c.notes
        resp.content_type = "text/html"
        resp.body = bookmark_form2.render(d=d,f=f,g=g,jsob=uid)
        resp.status=falcon.HTTP_301
        return
    def on_post(self, req, resp,uid,aa):
        details=req.params
        print details
        
        c=Bookmark_db.objects.get(u_id2=uid,location="http://"+aa)
        c.update(name=details.get('Name'),labels=details.get('Label'),notes=details.get('Notes'))
        print "updated successfully"
        resp.set_header('Location','/views/'+uid)
        resp.status=falcon.HTTP_301
        return

class delete(object):
    def on_get(self, req, resp,uid):
        if not Bookmark_db.objects(u_id2=uid):
            resp.content_type = "text/html"
            resp.status = falcon.HTTP_200
            temp="Nothing to delete"
            resp.body = bookmark1_template.render(temp=temp,jsob=uid)
            return
        for i in Bookmark_db.objects(u_id2=uid):
            i.delete()
        
        
        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        temp="files deleted successfully"
        resp.body = bookmark1_template.render(jsob=uid,temp=temp)
        return
class export(object):
    def on_get(self, req, resp,uid):
        list_url=[]
        if not Bookmark_db.objects(u_id2=uid):
            resp.set_header('Location','/views/'+uid)
            resp.status=falcon.HTTP_301
            return

        for i in Bookmark_db.objects(u_id2=uid):
            list_url.append(i.location)
            #print"rferf"
            #print list1
        
        
            #print 'dwed'
            
        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        resp.body = bookmark_view2.render(jsob=uid ,url_list=list_url)
        return 
    def on_post(self, req, resp,uid):

        url_list=req.params
        print url_list
        url1=[]
        print url_list.get('url')
        print "type", type(url_list.get('url'))
        if (type(url_list.get('url'))==list):

            url1=url_list.get('url')
            print url1
        else:
            url1.append(url_list.get('url'))

        wb=openpyxl.Workbook()
        sheet=wb.get_active_sheet()

        if not url1:
            
            resp.content_type = "text/html"
            resp.status = falcon.HTTP_200
            temp="Please select atleast one option"
            resp.body = bookmark_view2.render(jsob=uid ,temp=temp)
            return
        
        for j in url1:
            

            for i in Bookmark_db.objects(u_id2=uid,location=j):
                print "keys wala"    

                li=[]
                for f in i._data.keys():
                    if f=='id':
                        pass
                    else:

                        li.append(f)
                    print '1'
                print '2'
                    
                #print "listkey",li
        
                for k in range(1,len(li)):
            
                    sheet.cell(row=1,column=k).value=li[k-1]
                    print "keys done in excel"
                print '3'

            #print sheet
                row_num=2
                for i in url1:
                    print "url_list",i

                    for j in Bookmark_db.objects(u_id2=uid,location=i):
                        print "value",j._data.values()
                        li1=[]
                        
                        for k in range(0,len(j._data.values())):
                            if k==3:
                                pass
                            else:
                                li1.append(j._data.values()[k])

                    
                        print "li1",li1
                        for l in range(1,len(li1)):
                            sheet.cell(row=row_num,column=l).value=li1[l-1]
                            print "values done"
                        row_num+=1
                wb.save(UPLOAD_FOLDER+'/'+uid+'.xlsx')
                print 'd'
                resp.set_header('Location','/bookmark1/'+uid)
                resp.status=falcon.HTTP_301
                return
            #for l in range(1,len(li1)):


class import_file(object):
    def on_get(self, req, resp,uid):

        resp.content_type = "text/html"
        resp.status = falcon.HTTP_200
        resp.body = upload_file.render()
        
        
    def on_post(self,req, resp,uid):

        file1=req.params
        print file1.get('file_name')    
            

        wb = openpyxl.load_workbook('/home/sumit/bookmark_falcon/'+str(file1.get('file_name')))
        print wb.get_sheet_names()
        sheet=wb.get_active_sheet()
        d=sheet.get_highest_row()
        for i in range(2,d+1):
            try:
                if (Bookmark_db.objects.get(u_id2=uid,location=sheet.cell(row =i,column=4).value)):
                    continue
            except DoesNotExist:

                bookmark_db=Bookmark_db(name=sheet.cell(row =i,column=1).value,location=sheet.cell(row =i,column=4).value,
                    labels=sheet.cell(row =i,column=3).value,notes=sheet.cell(row =i,column=2).value,u_id2=uid).save()
        resp.set_header('Location','/views/'+uid)
        resp.status=falcon.HTTP_301
        return



class logout(object):
    def on_get(self, req, resp,uid):
        resp.content_type = "text/html"
        
        resp.body = Welcome_Template.render()
        resp.status = falcon.HTTP_301
        

    
